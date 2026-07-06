from datetime import datetime, date
import yaml
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from storage.database import get_relevant_jobs, get_all_jobs

COLUMN_MAP = {
    "title":               "Job Title",
    "company":             "Company",
    "location":            "Location",
    "source":              "Platform",
    "relevance_score":     "Match Score",
    "internship_friendly": "Intern Friendly",
    "experience_required": "Exp Required",
    "match_reason":        "Why Matched",
    "days_old":            "Days Old",
    "date_posted":         "Date Posted",
    "hr_email":            "HR Email 1",
    "hr_email_2":          "HR Email 2",
    "hr_email_3":          "HR Email 3",
    "hr_name":             "HR Name",
    "phone":               "Phone",
    "contact_form_url":    "Contact Form",
    "draft_email":         "Draft Email",
    "job_url":             "Apply Link",
    "status":              "Status",
}

SCORE_COLORS = {
    (9, 10): "C6EFCE",   # green
    (7,  8): "FFEB9C",   # yellow
    (0,  6): "FFC7CE",   # red
}

INTERN_FILL    = PatternFill("solid", fgColor="BDD7EE")   # blue — internship friendly
HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT    = Font(bold=True, color="FFFFFF")
TODAY_BORDER   = Border(
    left=Side(style="medium", color="FF0000"),
    right=Side(style="medium", color="FF0000"),
    top=Side(style="medium", color="FF0000"),
    bottom=Side(style="medium", color="FF0000"),
)


def _score_fill(score):
    try:
        s = int(score)
    except (TypeError, ValueError):
        return None
    for (lo, hi), color in SCORE_COLORS.items():
        if lo <= s <= hi:
            return PatternFill("solid", fgColor=color)
    return None


def _days_label(days_old):
    if days_old is None:
        return "Unknown"
    if days_old == 0:
        return "Today"
    if days_old == 1:
        return "Yesterday"
    return f"{days_old} days ago"


def _compute_days_old(date_posted_str):
    if not date_posted_str or str(date_posted_str).strip() in ("", "nan", "None"):
        return None
    try:
        dt = datetime.strptime(str(date_posted_str).strip()[:10], "%Y-%m-%d").date()
        return (date.today() - dt).days
    except Exception:
        return None


def _prepare_df(jobs):
    if not jobs:
        return pd.DataFrame()

    df = pd.DataFrame(jobs)

    # Compute days_old from date_posted if missing
    if "days_old" not in df.columns or df["days_old"].isna().all():
        df["days_old"] = df.get("date_posted", pd.Series()).apply(_compute_days_old)

    # Human-readable "Posted" column
    df["days_old_label"] = df["days_old"].apply(_days_label)

    # Boolean → readable
    if "internship_friendly" in df.columns:
        df["internship_friendly"] = df["internship_friendly"].apply(
            lambda x: "YES" if x else ""
        )

    cols = [c for c in COLUMN_MAP if c in df.columns]
    df = df[cols].rename(columns=COLUMN_MAP)

    # Sort: recency first (None last), then score
    score_col = "Match Score"
    if "Days Old" in df.columns:
        df = df.sort_values(
            ["Days Old", score_col],
            ascending=[True, False],
            na_position="last"
        )

    return df


def _write_sheet(ws, df):
    # Write data
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Style header
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    score_col_idx = None
    intern_col_idx = None
    days_col_idx = None
    for i, col in enumerate(df.columns, 1):
        if col == "Match Score":
            score_col_idx = i
        if col == "Intern Friendly":
            intern_col_idx = i
        if col == "Days Old":
            days_col_idx = i

    # Row styles
    for r_idx in range(2, ws.max_row + 1):
        score_val = ws.cell(row=r_idx, column=score_col_idx).value if score_col_idx else None
        intern_val = ws.cell(row=r_idx, column=intern_col_idx).value if intern_col_idx else None
        days_val = ws.cell(row=r_idx, column=days_col_idx).value if days_col_idx else None

        base_fill = _score_fill(score_val)

        # Internship-friendly overrides with blue
        row_fill = INTERN_FILL if intern_val == "YES" else base_fill

        for c_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if row_fill:
                cell.fill = row_fill
            # Red border for jobs posted today
            if days_val == 0:
                cell.border = TODAY_BORDER

    # Auto column widths
    for i, col in enumerate(df.columns, 1):
        max_len = max(df[col].astype(str).str.len().max(), len(col)) + 4
        ws.column_dimensions[get_column_letter(i)].width = min(max_len, 60)


def export_to_excel(output_path=None, min_score=7, include_all=False):
    with open("config.yaml") as f:
        config = yaml.safe_load(f)

    if output_path is None:
        output_path = config["storage"]["excel_path"]

    jobs = get_all_jobs() if include_all else get_relevant_jobs(min_score)

    if not jobs:
        print("No jobs to export.")
        return

    df_all = _prepare_df(jobs)

    # Last 24 hours sheet — include jobs with no date_posted but scraped today
    today_str = datetime.today().strftime("%Y-%m-%d")
    def _is_recent(row):
        days = row.get("Days Old") if isinstance(row, dict) else None
        if isinstance(days, (int, float)) and days <= 1:
            return True
        scraped = str(row.get("date_scraped", "") if isinstance(row, dict) else "")
        return scraped.startswith(today_str)
    if "Days Old" in df_all.columns:
        df_24h = df_all[df_all["Days Old"].apply(lambda x: isinstance(x, (int, float)) and x <= 1)]
    else:
        df_24h = pd.DataFrame()

    # Internship-friendly sheet
    df_intern = df_all[df_all.get("Intern Friendly", pd.Series()) == "YES"] \
        if "Intern Friendly" in df_all.columns else pd.DataFrame()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: All relevant jobs
        df_all.to_excel(writer, index=False, sheet_name="All Matches")
        _write_sheet(writer.sheets["All Matches"], df_all)

        # Sheet 2: Last 24 hours
        if not df_24h.empty:
            df_24h.to_excel(writer, index=False, sheet_name="Last 24hrs")
            _write_sheet(writer.sheets["Last 24hrs"], df_24h)

        # Sheet 3: Internship-friendly
        if not df_intern.empty:
            df_intern.to_excel(writer, index=False, sheet_name="Internship Friendly")
            _write_sheet(writer.sheets["Internship Friendly"], df_intern)

    sheets_info = f"All Matches ({len(df_all)})"
    if not df_24h.empty:
        sheets_info += f" | Last 24hrs ({len(df_24h)})"
    if not df_intern.empty:
        sheets_info += f" | Intern Friendly ({len(df_intern)})"

    print(f"Exported to {output_path} — Sheets: {sheets_info}")
    return output_path
