"""
Export a clean, filtered Excel sheet of fresher/junior jobs for manual applying.
Run anytime after the pipeline: python export_fresher_jobs.py

Filters applied:
  - Last 7 days only (+ null-date jobs from recent runs like HackerNews)
  - Score >= 7
  - No senior/lead/principal titles
  - No irrelevant tech titles (Android, Flutter, DevOps, etc.)
  - No jobs where description says 2+ years experience required
  - Yellow highlight = no description stored (verify exp before applying)

Output: fresher_jobs_share.xlsx
"""
import os
import re
import sqlite3
import sys
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from agent.exp_filter import has_experience_requirement

# ── Filters ────────────────────────────────────────────────────────────────────
SENIOR_RE = re.compile(
    r"\b(senior|sr\.?\s|lead\s|principal|staff\s+eng|engineering\s+manager|"
    r"director|head\s+of|vice\s+pres|vp\s+of|architect(?!\s+as))\b", re.IGNORECASE)
IRRELEVANT_RE = re.compile(
    r"\b(android|flutter|ios\s+dev|swift\s+dev|kotlin|"
    r"devops|sre\s+|site\s+reliability|data\s+scientist|data\s+engineer|"
    r"machine\s+learning\s+eng|pyspark|salesforce|sap\s+|"
    r"manufacturing|quality\s+assurance|qa\s+eng|guard\b|"
    r"transportation\s+rep|relationship\s+manager)\b", re.IGNORECASE)
EXP_FIELD_RE = re.compile(
    r"(2\+|3\+|4\+|5\+|6\+|7\+|8\+|9\+|10\+|several|minimum\s+[3-9]|[3-9]\s+year)",
    re.IGNORECASE)


def export(days=7, output="fresher_jobs_share.xlsx"):
    cutoff = (datetime.today() - timedelta(days=days)).strftime("%Y-%m-%d")

    conn = sqlite3.connect("jobs.db")
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT title, company, location, job_url, relevance_score,
               hr_email, hr_email_2, hr_email_3, phone, date_posted,
               experience_required, description
        FROM jobs
        WHERE relevance_score >= 7
          AND (date_posted IS NULL OR date_posted = 'nan' OR date_posted >= ?)
        ORDER BY date_posted DESC, relevance_score DESC
    """, (cutoff,)).fetchall()
    conn.close()

    filtered = []
    stats = dict(senior=0, irrelevant=0, exp_field=0, exp_desc=0)
    for r in rows:
        title = str(r["title"] or "")
        exp   = str(r["experience_required"] or "")
        desc  = str(r["description"] or "")
        if SENIOR_RE.search(title):                             stats["senior"] += 1;     continue
        if IRRELEVANT_RE.search(title):                        stats["irrelevant"] += 1; continue
        if EXP_FIELD_RE.search(exp):                           stats["exp_field"] += 1;  continue
        if has_experience_requirement(desc):                      stats["exp_desc"] += 1;   continue
        no_desc = desc.strip() in ("nan", "", "None")
        filtered.append((r, no_desc))

    total_removed = sum(stats.values())
    print(f"  From DB  : {len(rows)} jobs (last {days} days)")
    print(f"  Filtered : {total_removed} removed "
          f"(senior={stats['senior']} irrelevant={stats['irrelevant']} "
          f"2+yrs_field={stats['exp_field']} 2+yrs_desc={stats['exp_desc']})")
    print(f"  Output   : {len(filtered)} clean fresher jobs "
          f"({sum(1 for _, nd in filtered if nd)} yellow = no description)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fresher Jobs"

    headers = ["#", "Job Title", "Company", "Location", "Score",
               "Date Posted", "Apply", "HR Email 1", "HR Email 2", "HR Email 3", "Phone"]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    warn_fill   = PatternFill("solid", fgColor="FFF2CC")  # yellow — no description, verify exp
    normal_fill = PatternFill("solid", fgColor="EBF3FB")  # blue alternate rows

    for i, (row, no_desc) in enumerate(filtered, 1):
        ws.cell(row=i + 1, column=1, value=i)
        title_val = row["title"] + (" [CHECK EXP]" if no_desc else "")
        title_cell = ws.cell(row=i + 1, column=2, value=title_val)
        if no_desc:
            title_cell.font = Font(color="7F6000")
        ws.cell(row=i + 1, column=3, value=row["company"])
        ws.cell(row=i + 1, column=4, value=row["location"])
        ws.cell(row=i + 1, column=5, value=row["relevance_score"])
        d = str(row["date_posted"] or "")
        ws.cell(row=i + 1, column=6, value="" if d in ("nan", "None", "") else d)
        url = row["job_url"]
        if url:
            cell = ws.cell(row=i + 1, column=7, value="Apply")
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
        else:
            ws.cell(row=i + 1, column=7, value="")
        ws.cell(row=i + 1, column=8,  value=row["hr_email"]   or "")
        ws.cell(row=i + 1, column=9,  value=row["hr_email_2"] or "")
        ws.cell(row=i + 1, column=10, value=row["hr_email_3"] or "")
        ws.cell(row=i + 1, column=11, value=row["phone"]      or "")

        row_fill = warn_fill if no_desc else (normal_fill if i % 2 == 0 else None)
        if row_fill:
            for col in range(1, 12):
                ws.cell(row=i + 1, column=col).fill = row_fill

    for col, w in enumerate([4, 45, 25, 18, 7, 12, 8, 28, 28, 28, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(filtered) + 1}"

    wb.save(output)
    print(f"  Saved    : {output}")


if __name__ == "__main__":
    print("\n" + "=" * 55)
    print("  JobPilot AI — Fresher Jobs Export")
    print("=" * 55 + "\n")
    export()
    print()
