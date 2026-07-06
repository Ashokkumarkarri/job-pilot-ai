"""
Send notification email with this run's matched jobs as an attached Excel sheet.
"""
import os
import sys
import smtplib
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from agent.exp_filter import has_experience_requirement, SENIOR_TITLE_RE, IRRELEVANT_TITLE_RE

GMAIL_ADDRESS  = os.getenv("GMAIL_ADDRESS", "")
GMAIL_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")
SMTP_HOST      = "smtp.gmail.com"
SMTP_PORT      = 587


def _build_sheet(jobs: list, out_path: str):
    filtered = []
    for j in jobs:
        title = str(j.get("title") or "")
        desc  = str(j.get("description") or "")
        if SENIOR_TITLE_RE.search(title) or IRRELEVANT_TITLE_RE.search(title):
            continue
        if has_experience_requirement(desc):
            continue
        filtered.append(j)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "New Jobs"

    headers = ["#", "Platform", "Job Title", "Company", "Location", "Score", "Apply"]
    hfill   = PatternFill("solid", fgColor="1F4E79")
    hfont   = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    alt_fill = PatternFill("solid", fgColor="EBF3FB")

    for i, job in enumerate(filtered, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=str(job.get("source", "")).title())
        ws.cell(row=i + 1, column=3, value=job.get("title", ""))
        ws.cell(row=i + 1, column=4, value=job.get("company", ""))
        ws.cell(row=i + 1, column=5, value=job.get("location", ""))
        ws.cell(row=i + 1, column=6, value=job.get("relevance_score", ""))
        url = job.get("job_url", "")
        if url:
            cell = ws.cell(row=i + 1, column=7, value="Apply")
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
        else:
            ws.cell(row=i + 1, column=7, value="")
        if i % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=i + 1, column=col).fill = alt_fill

    for col, w in enumerate([4, 12, 40, 25, 18, 7, 8], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(filtered) + 1}"

    wb.save(out_path)
    return len(filtered)


def send_report(jobs: list, recipient: str = None, tag: str = "[APPLY NOW]") -> bool:
    if not GMAIL_ADDRESS or not GMAIL_PASSWORD:
        print("  Notify: Gmail credentials not configured — skipping.")
        return False

    to_addr    = recipient or GMAIL_ADDRESS
    now        = datetime.now().strftime("%d %b %Y %I:%M %p")
    time_str   = datetime.now().strftime("%I:%M %p")
    sheet_path = f"run_jobs_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx"

    job_count = _build_sheet(jobs, sheet_path)
    subject   = f"JobPilot | {time_str} — {job_count} jobs {tag}"

    body = f"""\
Hi Kumar,

JobPilot AI completed a run. Here's what was found:

  Matched jobs this run : {len(jobs)}
  After fresher filter  : {job_count}
  Time                  : {now}

Sheet attached — all Apply links are clickable.

— JobPilot AI
"""

    msg = MIMEMultipart("mixed")
    msg["From"]    = GMAIL_ADDRESS
    msg["To"]      = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    if os.path.exists(sheet_path):
        with open(sheet_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=sheet_path)
        msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(GMAIL_ADDRESS, GMAIL_PASSWORD)
            server.sendmail(GMAIL_ADDRESS, to_addr, msg.as_bytes())
        try:
            os.remove(sheet_path)
        except Exception:
            pass
        return True
    except Exception as e:
        print(f"  Notify: Failed to send — {e}")
        return False
